"""
generate_clv_report.py
Generates the Project 02 CLV Segmentation Excel report.
Run from the project root:
    python reports/generate_clv_report.py
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paths
ROOT        = Path(__file__).resolve().parent.parent
PROCESSED   = ROOT / 'data' / 'processed'
OUTPUT_PATH = ROOT / 'reports' / 'excel_outputs' / 'clv_segmentation_report.xlsx'
OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

#  Load all processed data 
print('Loading processed data...')
rfm         = pd.read_csv(PROCESSED / 'rfm_segments.csv')
clv         = pd.read_csv(PROCESSED / 'clv_predictions.csv')
channel_roi = pd.read_csv(PROCESSED / 'channel_roi_summary.csv', index_col=0)
regional    = pd.read_csv(PROCESSED / 'regional_clv_summary.csv', index_col=0)
country     = pd.read_csv(PROCESSED / 'country_clv_summary.csv', index_col=0)

print(f'  RFM segments:     {len(rfm):,} customers')
print(f'  CLV predictions:  {len(clv):,} customers')
print(f'  Channel summary:  {len(channel_roi)} channels')
print(f'  Regional summary: {len(regional)} regions')
print(f'  Country summary:  {len(country)} countries')

#  Style constants 
DARK_BLUE    = '1F4E79'
MID_BLUE     = '2E75B6'
LIGHT_BLUE   = 'D6E4F0'
WHITE        = 'FFFFFF'
DARK_GREY    = '404040'
LIGHT_GREY   = 'F2F2F2'
GREEN        = '1E6B3C'
LIGHT_GREEN  = 'E2EFDA'
RED          = 'C00000'
LIGHT_RED    = 'FCE4D6'
AMBER        = 'ED7D31'
LIGHT_AMBER  = 'FFF2CC'
FONT         = 'Arial'

thin       = Side(style='thin',   color='CCCCCC')
med        = Side(style='medium', color=MID_BLUE)
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(size=10, bold=True, color=WHITE):
    return Font(name=FONT, size=size, bold=bold, color=color)

def body(size=10, bold=False, color=DARK_GREY):
    return Font(name=FONT, size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def write_header_row(ws, row, headers, col_widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = hdr(10)
        c.fill = fill(MID_BLUE)
        c.alignment = center()
        c.border = THIN_BORDER
        if col_widths and i <= len(col_widths):
            ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]

def write_data_row(ws, row, values, alt=False, bold=False):
    bg = LIGHT_GREY if alt else WHITE
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = body(bold=bold)
        c.fill = fill(bg)
        c.alignment = left()
        c.border = THIN_BORDER

def title_block(ws, row, text, subtitle=None):
    ws.merge_cells(f'A{row}:{get_column_letter(8)}{row}')
    c = ws[f'A{row}']
    c.value = text
    c.font = Font(name=FONT, size=14, bold=True, color=DARK_BLUE)
    c.fill = fill(LIGHT_BLUE)
    c.alignment = left()
    ws.row_dimensions[row].height = 30
    if subtitle:
        ws.merge_cells(f'A{row+1}:{get_column_letter(8)}{row+1}')
        s = ws[f'A{row+1}']
        s.value = subtitle
        s.font = Font(name=FONT, size=9, color='666666', italic=True)
        s.alignment = left()
        ws.row_dimensions[row+1].height = 16

def section_label(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = Font(
        name=FONT, size=11, bold=True, color=DARK_BLUE)
    ws.row_dimensions[row].height = 22

wb = Workbook()

# 
# SHEET 1 — EXECUTIVE SUMMARY
# 
ws1 = wb.create_sheet('01 Executive Summary', 0)
wb.remove(wb['Sheet'])
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 6

# Title banner
for col in range(1, 9):
    ws1.cell(row=2, column=col).fill = fill(DARK_BLUE)
ws1.merge_cells('A2:H2')
t = ws1['A2']
t.value = 'GameZone Analytics — Project 02 CLV Segmentation Report'
t.font = Font(name=FONT, size=16, bold=True, color=WHITE)
t.alignment = center()
ws1.row_dimensions[2].height = 36

ws1.merge_cells('A3:H3')
s = ws1['A3']
s.value = f'Generated: {datetime.now().strftime("%d %B %Y")}   |   {len(rfm):,} customers analysed   |   4 notebooks'
s.font = Font(name=FONT, size=9, color='888888')
s.alignment = center()
s.fill = fill(LIGHT_GREY)
ws1.row_dimensions[3].height = 18
ws1.row_dimensions[4].height = 10

# Key metrics
seg_counts  = rfm['segment'].value_counts()
total_rev   = rfm['monetary'].sum()
champ_rev   = rfm[rfm['segment'] == 'Champions']['monetary'].sum()
loyal_rev   = rfm[rfm['segment'] == 'Loyal Customers']['monetary'].sum()
top_channel = channel_roi['avg_total_spend'].idxmax()
top_country = country['avg_spend'].idxmax()

metrics = [
    ('Total Customers',         f"{len(rfm):,}",                       DARK_BLUE),
    ('Champions',               f"{seg_counts.get('Champions',0):,}",  MID_BLUE),
    ('Loyal Customers',         f"{seg_counts.get('Loyal Customers',0):,}", MID_BLUE),
    ('At Risk',                 f"{seg_counts.get('At Risk',0):,}",    AMBER),
    ('Best Channel (CLV)',      top_channel,                           GREEN),
    ('Best Country (Spend)',    top_country,                           GREEN),
    ('Loyal Rev Share',         f"{loyal_rev/total_rev*100:.0f}%",     GREEN),
    ('Champion Rev Share',      f"{champ_rev/total_rev*100:.0f}%",     MID_BLUE),
]

for i, (label, value, color) in enumerate(metrics):
    col = i + 1
    lbl = ws1.cell(row=5, column=col, value=label)
    lbl.font = Font(name=FONT, size=8, bold=True, color=WHITE)
    lbl.fill = fill(MID_BLUE)
    lbl.alignment = center()
    lbl.border = THIN_BORDER
    ws1.row_dimensions[5].height = 18

    val = ws1.cell(row=6, column=col, value=value)
    val.font = Font(name=FONT, size=14, bold=True, color=color)
    val.fill = fill(LIGHT_BLUE)
    val.alignment = center()
    val.border = THIN_BORDER
    ws1.row_dimensions[6].height = 32

ws1.row_dimensions[7].height = 10

# Key findings table
section_label(ws1, 8, 'Key Findings Summary')
write_header_row(ws1, 9,
    ['Notebook', 'Finding', 'Key Metric', 'Implication'],
    col_widths=[24, 50, 20, 36])

findings = [
    ('NB1 — RFM',       'Loyal Customers (6.9% of base) drive 36% of revenue',
     '36% rev / 6.9% base', 'Protect Loyal segment — highest ROI per customer'),
    ('NB1 — RFM',       'At Risk segment: 43.6% of customers, 27.5% of revenue',
     '8,594 customers', 'Largest win-back opportunity by both count and revenue'),
    ('NB1 — RFM',       'Champions are frequency-driven ($435 avg), Loyal are spend-driven ($1,613)',
     '$1,613 vs $435', 'Different upsell strategies needed for each segment'),
    ('NB2 — BG/NBD',    '98.9% of customers purchased only once',
     '98.9% one-time', 'Dataset too sparse for BG/NBD signal — RFM is primary CLV tool'),
    ('NB2 — BG/NBD',    'Loyal Customers: 98.4% probability still active',
     '98.4% prob alive', 'Retention focus — these customers have not churned'),
    ('NB3 — Channel',   'Affiliate has highest avg spend per customer ($343)',
     '$343 avg / 16.8% HV', 'Scale affiliate investment — same quality as direct at lower volume'),
    ('NB3 — Channel',   'Email underperforms: 15.1% customers, only 10% revenue',
     '0.66x ROI ratio', 'Restructure email — currently attracts price-sensitive buyers'),
    ('NB3 — Channel',   'Channel is stronger CLV predictor than geography (68% vs 8% gap)',
     'Channel > Geography', 'Optimise channel mix before geographic targeting'),
    ('NB4 — Geo',       'Japan: highest avg spend ($472) and best ROI ratio (1.52x)',
     '$472 / 1.52x ROI', 'Japan is priority retention and growth market'),
    ('NB4 — Geo',       'AU, CA, IE below break-even on ROI ratio',
     '0.73-0.88x ROI', 'Review acquisition spend in English-speaking markets'),
    ('NB4 — Geo',       'Segment mix is uniform across all regions (no geographic segmentation signal)',
     'All regions ~9% Champ', 'Do not build region-specific segments — RFM is universal'),
]

severity_map = {
    'NB1 — RFM':    (LIGHT_BLUE,  MID_BLUE),
    'NB2 — BG/NBD': (LIGHT_GREY,  DARK_GREY),
    'NB3 — Channel':(LIGHT_GREEN, GREEN),
    'NB4 — Geo':    (LIGHT_AMBER, AMBER),
}

for i, (nb, finding, metric, impl) in enumerate(findings):
    row = 10 + i
    alt = i % 2 == 0
    bg, fg = severity_map.get(nb, (LIGHT_GREY, DARK_GREY))
    write_data_row(ws1, row, [nb, finding, metric, impl], alt=alt)
    nb_cell = ws1.cell(row=row, column=1)
    nb_cell.fill = fill(bg)
    nb_cell.font = Font(name=FONT, size=9, bold=True, color=fg)
    nb_cell.alignment = center()
    ws1.row_dimensions[row].height = 28

ws1.freeze_panes = 'A9'
for col_letter, width in zip('ABCDEFGH', [24, 50, 20, 36, 12, 12, 12, 12]):
    ws1.column_dimensions[col_letter].width = width

# 
# SHEET 2 — RFM SEGMENTS
# 
ws2 = wb.create_sheet('02 RFM Segments')
ws2.sheet_view.showGridLines = False

title_block(ws2, 1, 'RFM Segmentation — Notebook 1',
    'K-Means clustering (k=4) on StandardScaler-normalised Recency, Frequency, Monetary')

section_label(ws2, 4, 'Segment Profiles')
write_header_row(ws2, 5,
    ['Segment', 'Customers', '% of Base', 'Avg Recency (days)',
     'Avg Orders', 'Avg Spend ($)', 'Total Revenue ($)', 'Revenue Share (%)'],
    col_widths=[20, 13, 13, 20, 13, 16, 18, 18])

seg_profile = rfm.groupby('segment').agg(
    customers     = ('USER_ID',    'count'),
    avg_recency   = ('recency',    'mean'),
    avg_frequency = ('frequency',  'mean'),
    avg_monetary  = ('monetary',   'mean'),
    total_revenue = ('monetary',   'sum')
).round(2)
seg_profile['pct_base']    = (seg_profile['customers'] / len(rfm) * 100).round(1)
seg_profile['rev_share']   = (seg_profile['total_revenue'] / rfm['monetary'].sum() * 100).round(1)

seg_order  = ['Champions', 'Loyal Customers', 'At Risk', 'Lapsed']
seg_colors = {
    'Champions':      (LIGHT_BLUE,  MID_BLUE),
    'Loyal Customers':(LIGHT_BLUE,  MID_BLUE),
    'At Risk':        (LIGHT_AMBER, AMBER),
    'Lapsed':         (LIGHT_RED,   RED),
}

for i, seg in enumerate(seg_order):
    if seg not in seg_profile.index:
        continue
    row  = 6 + i
    r    = seg_profile.loc[seg]
    bg, fg = seg_colors.get(seg, (WHITE, DARK_GREY))
    vals = [seg, int(r['customers']), f"{r['pct_base']:.1f}%", # pyright: ignore[reportArgumentType]
            f"{r['avg_recency']:.0f}", f"{r['avg_frequency']:.2f}",
            f"${r['avg_monetary']:,.2f}", f"${r['total_revenue']:,.0f}",
            f"{r['rev_share']:.1f}%"]
    write_data_row(ws2, row, vals)
    for col in range(1, 9):
        c = ws2.cell(row=row, column=col)
        c.fill = fill(bg)
        c.font = Font(name=FONT, size=10, bold=(col == 1), color=fg if col == 1 else DARK_GREY)
    ws2.row_dimensions[row].height = 22

ws2.row_dimensions[10].height = 10
section_label(ws2, 11, 'Business Recommendations by Segment')
write_header_row(ws2, 12,
    ['Segment', 'Priority', 'Recommended Action'],
    col_widths=[20, 14, 60])

recs = [
    ('Champions',       'HIGH',   'Cross-sell high-ticket items — Switch buyers are PS5 prospects. No heavy discounting.'),
    ('Loyal Customers', 'CRITICAL','Premium personalised outreach. Bundle accessories. Never discount — they buy at full price.'),
    ('At Risk',         'HIGH',   'Time-sensitive win-back campaign. Single targeted incentive. 30-day response window.'),
    ('Lapsed',          'LOW',    'Single low-cost re-engagement email only. Suppress after 2 attempts with no response.'),
]

priority_colors = {'CRITICAL': (LIGHT_RED, RED), 'HIGH': (LIGHT_AMBER, AMBER), 'LOW': (LIGHT_GREY, DARK_GREY)}

for i, (seg, priority, rec) in enumerate(recs):
    row = 13 + i
    write_data_row(ws2, row, [seg, priority, rec], alt=i % 2 == 0)
    p_cell = ws2.cell(row=row, column=2)
    bg_p, fg_p = priority_colors.get(priority, (WHITE, DARK_GREY))
    p_cell.fill = fill(bg_p)
    p_cell.font = Font(name=FONT, size=10, bold=True, color=fg_p)
    p_cell.alignment = center()
    ws2.row_dimensions[row].height = 28

ws2.freeze_panes = 'A5'

# 
# SHEET 3 — CHANNEL ROI
# 
ws3 = wb.create_sheet('03 Channel ROI')
ws3.sheet_view.showGridLines = False

title_block(ws3, 1, 'Channel ROI Attribution — Notebook 3',
    'Kruskal-Wallis H=375.60, p<0.0001 — channel spend differences are statistically significant')

section_label(ws3, 4, 'Channel Performance Summary')
write_header_row(ws3, 5,
    ['Channel', 'Customers', 'Avg Spend ($)', 'Revenue Share (%)',
     'Customer Share (%)', 'High-Value Rate (%)', 'Avg RFM Score'],
    col_widths=[16, 13, 15, 18, 18, 20, 16])

ch_sorted = channel_roi.sort_values('avg_total_spend', ascending=False)
overall_avg_spend = channel_roi['avg_total_spend'].mean()

for i, (ch, row_data) in enumerate(ch_sorted.iterrows()):
    row = 6 + i
    alt = i % 2 == 0
    avg_sp = row_data.get('avg_total_spend', 0)
    write_data_row(ws3, row, [
        ch,
        f"{int(row_data.get('customers', 0)):,}",
        f"${avg_sp:,.2f}",
        f"{row_data.get('revenue_share_pct', 0):.1f}%",
        f"{row_data.get('order_share_pct', 0):.1f}%",
        f"{row_data.get('high_value_customer_pct', 0):.1f}%",
        f"{row_data.get('avg_rfm_score', 0):.2f}"
    ], alt=alt)
    spend_cell = ws3.cell(row=row, column=3)
    if avg_sp > overall_avg_spend:
        spend_cell.fill = fill(LIGHT_GREEN)
        spend_cell.font = Font(name=FONT, size=10, bold=True, color=GREEN)
    else:
        spend_cell.fill = fill(LIGHT_RED)
        spend_cell.font = Font(name=FONT, size=10, bold=True, color=RED)
    ws3.row_dimensions[row].height = 20

ws3.row_dimensions[6 + len(ch_sorted) + 1].height = 10
offset3 = 6 + len(ch_sorted) + 2
section_label(ws3, offset3, 'Statistical Test Results')
write_header_row(ws3, offset3+1,
    ['Test', 'Result', 'Interpretation'],
    col_widths=[28, 20, 52])

tests = [
    ('Kruskal-Wallis (all channels)', 'H=375.60, p<0.0001', 'Spend differences across channels are statistically real'),
    ('Affiliate vs Direct',           'p=0.122 (NS)',        'Not significantly different — same customer quality'),
    ('Affiliate vs Email',            'p<0.0001 ',         'Affiliate customers significantly higher spend than email'),
    ('Direct vs Email',               'p<0.0001 ',         'Direct customers significantly higher spend than email'),
    ('Direct vs Social Media',        'p=0.004 ',          'Direct customers significantly higher spend than social media'),
    ('Email vs Social Media',         'p<0.0001 ',         'Email and social media both below average but differ from each other'),
]

for i, (test, result, interp) in enumerate(tests):
    row = offset3 + 2 + i
    write_data_row(ws3, row, [test, result, interp], alt=i % 2 == 0)
    ws3.row_dimensions[row].height = 20

ws3.freeze_panes = 'A5'

# 
# SHEET 4 — GEOGRAPHIC CLV
# 
ws4 = wb.create_sheet('04 Geographic CLV')
ws4.sheet_view.showGridLines = False

title_block(ws4, 1, 'Geographic CLV Analysis — Notebook 4',
    'Bootstrap 95% CIs — Kruskal-Wallis H=121.12, p<0.0001 — regional differences statistically significant')

section_label(ws4, 4, 'Regional CLV Summary')
write_header_row(ws4, 5,
    ['Region', 'Customers', 'Avg Spend ($)', 'Revenue Share (%)',
     'Customer Share (%)', 'High-Value Rate (%)', 'CI Lower ($)', 'CI Upper ($)'],
    col_widths=[14, 13, 15, 18, 18, 20, 14, 14])

reg_sorted = regional.sort_values('avg_spend', ascending=False)
overall_reg_avg = regional['avg_spend'].mean()

for i, (reg, r) in enumerate(reg_sorted.iterrows()):
    row = 6 + i
    alt = i % 2 == 0
    write_data_row(ws4, row, [
        reg,
        f"{int(r.get('customers', 0)):,}",
        f"${r.get('avg_spend', 0):,.2f}",
        f"{r.get('revenue_share_pct', 0):.1f}%",
        f"{r.get('customer_share_pct', 0):.1f}%",
        f"{r.get('high_value_customer_pct', 0):.1f}%",
        '—', '—'
    ], alt=alt)
    ws4.row_dimensions[row].height = 20

ws4.row_dimensions[6 + len(reg_sorted) + 1].height = 10
offset4 = 6 + len(reg_sorted) + 2
section_label(ws4, offset4, 'Top 20 Countries — CLV Rankings')
write_header_row(ws4, offset4+1,
    ['Country', 'Customers', 'Avg Spend ($)', 'Median Spend ($)',
     'Revenue Share (%)', 'Customer Share (%)', 'ROI Ratio'],
    col_widths=[12, 13, 15, 16, 18, 18, 13])

country_sorted = country.sort_values('avg_spend', ascending=False)
overall_country_avg = country['avg_spend'].mean()

for i, (cc, r) in enumerate(country_sorted.iterrows()):
    row = offset4 + 2 + i
    alt = i % 2 == 0
    roi = r.get('roi_ratio', 1.0)
    write_data_row(ws4, row, [
        cc,
        f"{int(r.get('customers', 0)):,}",
        f"${r.get('avg_spend', 0):,.2f}",
        f"${r.get('median_spend', 0):,.2f}",
        f"{r.get('revenue_share_pct', 0):.1f}%",
        f"{r.get('customer_share_pct', 0):.1f}%",
        f"{roi:.2f}x"
    ], alt=alt)
    roi_cell = ws4.cell(row=row, column=7)
    if roi > 1.0:
        roi_cell.fill = fill(LIGHT_GREEN)
        roi_cell.font = Font(name=FONT, size=10, bold=True, color=GREEN)
    else:
        roi_cell.fill = fill(LIGHT_RED)
        roi_cell.font = Font(name=FONT, size=10, bold=True, color=RED)
    avg_cell = ws4.cell(row=row, column=3)
    if r.get('avg_spend', 0) > overall_country_avg:
        avg_cell.fill = fill(LIGHT_GREEN)
        avg_cell.font = Font(name=FONT, size=10, color=GREEN)
    ws4.row_dimensions[row].height = 18

ws4.freeze_panes = 'A5'

# 
# SHEET 5 — MODEL NOTES (BG/NBD)
# 
ws5 = wb.create_sheet('05 BG-NBD Model Notes')
ws5.sheet_view.showGridLines = False

title_block(ws5, 1, 'BG/NBD CLV Model — Notebook 2',
    'Probabilistic CLV model — limitations documented for this dataset')

notes = [
    ('Model', 'BG/NBD (Beta-Geometric Negative Binomial Distribution) + Gamma-Gamma'),
    ('Library', 'lifetimes v0.11.3'),
    ('Customers modelled', f"{len(clv):,}"),
    ('Repeat buyers (frequency > 0)', '93 (1.1% of customers)'),
    ('One-time buyers (frequency = 0)', '8,733 (98.9% of customers)'),
    ('Prediction horizon', '12 weeks'),
    ('Model fit', 'Successful — parameters converged'),
    ('Validation', 'Predicted vs actual transaction counts closely matched'),
]

write_header_row(ws5, 4, ['Parameter', 'Value'], col_widths=[36, 52])
for i, (param, value) in enumerate(notes):
    row = 5 + i
    write_data_row(ws5, row, [param, value], alt=i % 2 == 0)
    ws5.row_dimensions[row].height = 20

ws5.row_dimensions[5 + len(notes) + 1].height = 10
section_label(ws5, 5 + len(notes) + 2, 'Why BG/NBD Predictions Are Near Zero')

limitations = [
    'GameZone sells high-ticket consumer electronics with 2-4 year repurchase cycles.',
    'The 2-year dataset (2019-2021) is insufficient to capture repeat purchase patterns.',
    '98.9% of customers made exactly one purchase — BG/NBD needs repeat transactions.',
    'The model fitted correctly and validated well — the near-zero CLV values accurately reflect the data.',
    'This is a product category constraint, not a modelling error.',
    'Use RFM segments (Notebook 1) as the primary CLV signal for this dataset.',
    'A 5+ year dataset or higher-frequency product category would enable full BG/NBD differentiation.',
]

lim_offset = 5 + len(notes) + 3
write_header_row(ws5, lim_offset, ['Limitation / Explanation'], col_widths=[88])
for i, lim in enumerate(limitations):
    row = lim_offset + 1 + i
    c = ws5.cell(row=row, column=1, value=f'• {lim}')
    c.font = body(size=10)
    c.fill = fill(LIGHT_GREY if i % 2 == 0 else WHITE)
    c.alignment = left()
    c.border = THIN_BORDER
    ws5.row_dimensions[row].height = 20

ws5.freeze_panes = 'A4'

# Save
wb.save(OUTPUT_PATH)
print(f'\nReport saved → {OUTPUT_PATH}')
print(f'   Sheets: {wb.sheetnames}')