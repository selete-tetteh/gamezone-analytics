"""

generate_data_quality_report.py

Generates the Project 01 Excel data quality report from processed CSV files.
Run from the project root:
    python reports/generate_data_quality_report.py      
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils.dataframe import dataframe_to_rows

# Define file paths
ROOT              = Path(__file__).resolve().parent.parent
PROCESSED         = ROOT / 'data' / 'processed'
OUTPUT_PATH       = ROOT / 'reports' / 'excel_outputs' / 'data_quality_report.xlsx'
OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

# Load all processed data
print('Loading processed data...')
df_master     = pd.read_csv(PROCESSED / 'orders_clean_master.csv',
                             parse_dates=['PURCHASE_TS', 'SHIP_TS'])
df_anomalous  = pd.read_csv(PROCESSED / 'anomalous_orders.csv',
                             parse_dates=['PURCHASE_TS', 'SHIP_TS'])
df_validation = pd.read_csv(PROCESSED / 'validation_results.csv')
df_name_map   = pd.read_csv(PROCESSED / 'product_name_mapping.csv')
df_id_map     = pd.read_csv(PROCESSED / 'product_id_mapping.csv')
print(f'  Master dataset:     {len(df_master):,} rows')
print(f'  Anomalous orders:   {len(df_anomalous):,} rows')
print(f'  Validation results: {len(df_validation):,} rows')


# Style Constants
DARK_BLUE   = '1F4E79'
MID_BLUE    = '2E75B6'
LIGHT_BLUE  = 'D6E4F0'
WHITE       = 'FFFFFF'
DARK_GREY   = '404040'
LIGHT_GREY  = 'F2F2F2'
MID_GREY    = 'D9D9D9'
GREEN       = '1E6B3C'
LIGHT_GREEN = 'E2EFDA'
RED         = 'C00000'
LIGHT_RED   = 'FCE4D6'
AMBER       = 'ED7D31'
LIGHT_AMBER = 'FCE4D6'
FONT_NAME   = 'Arial'
 
thin  = Side(style='thin',   color='CCCCCC')
med   = Side(style='medium', color=MID_BLUE)
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MED_BORDER  = Border(left=med,  right=med,  top=med,  bottom=med)
 
def hdr_font(size=11, bold=True, color=WHITE):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color)
 
def body_font(size=10, bold=False, color=DARK_GREY):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color)
 
def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)
 
def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)
 
def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)
 
def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width
 
def apply_header_row(ws, row_num, headers, col_widths=None):
    """Write a formatted header row with dark blue background."""
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=i, value=header)
        cell.font      = hdr_font(size=10)
        cell.fill      = fill(MID_BLUE)
        cell.alignment = center()
        cell.border    = THIN_BORDER
        if col_widths and i <= len(col_widths):
            ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]
 
def apply_data_row(ws, row_num, values, alt=False):
    """Write a formatted data row with alternating background."""
    bg = LIGHT_GREY if alt else WHITE
    for i, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=i, value=val)
        cell.font      = body_font()
        cell.fill      = fill(bg)
        cell.alignment = left()
        cell.border    = THIN_BORDER
 
def title_cell(ws, row, col, text, size=14):
    """Write a large title cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = Font(name=FONT_NAME, size=size, bold=True, color=DARK_BLUE)
    cell.alignment = left()
    return cell
 
def subtitle_cell(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = Font(name=FONT_NAME, size=10, color='666666', italic=True)
    cell.alignment = left()
    return cell
 
def metric_block(ws, row, col, label, value, value_color=DARK_BLUE):
    """Write a label + value metric pair."""
    lbl = ws.cell(row=row, column=col, value=label)
    lbl.font      = Font(name=FONT_NAME, size=9, color='888888')
    lbl.alignment = left()
    val = ws.cell(row=row+1, column=col, value=value)
    val.font      = Font(name=FONT_NAME, size=16, bold=True, color=value_color)
    val.alignment = left()
    return val
 
wb = Workbook()

# Sheet 1 - Executive Summary
ws1 = wb.active
assert ws1 is not None
ws1.title = '01 Executive Summary'
ws1.sheet_view.showGridLines = False # pyright: ignore[reportOptionalMemberAccess]
ws1.row_dimensions[1].height = 6 # pyright: ignore[reportOptionalMemberAccess]

# Title banner
for col in range(1, 8):
    cell = ws1.cell(row=2, column=col) # pyright: ignore[reportOptionalMemberAccess]
    cell.fill = fill(DARK_BLUE)
 
ws1.merge_cells('A2:G2')
title = ws1['A2']
title.value     = 'GameZone Analytics — Project 01 Data Quality Report'
title.font      = Font(name=FONT_NAME, size=16, bold=True, color=WHITE)
title.alignment = center()
ws1.row_dimensions[2].height = 36
 
ws1.merge_cells('A3:G3')
sub = ws1['A3']
sub.value     = f'Generated: {datetime.now().strftime("%d %B %Y")}   |   Dataset: GameZone Orders 2019–2021   |   Final row count: {len(df_master):,}'
sub.font      = Font(name=FONT_NAME, size=9, color='888888')
sub.alignment = center()
sub.fill      = fill(LIGHT_GREY)
ws1.row_dimensions[3].height = 18
 
ws1.row_dimensions[4].height = 12

# Key metrics row
anomaly_rate    = len(df_anomalous) / len(df_master) * 100
zero_price      = df_master['IS_ZERO_PRICE'].sum() if 'IS_ZERO_PRICE' in df_master.columns else 29
price_outliers  = df_master['IS_PRICE_OUTLIER'].sum() if 'IS_PRICE_OUTLIER' in df_master.columns else 1430
val_pass_rate   = df_validation['passed'].mean() * 100
 
metrics = [
    ('Total Orders',        f'{len(df_master):,}',      DARK_BLUE),
    ('Temporal Anomalies',  f'{len(df_anomalous):,}',   AMBER),
    ('$0 Price Orders',     f'{int(zero_price):,}',     AMBER),
    ('Price Outliers',      f'{int(price_outliers):,}', AMBER),
    ('Duplicate Orders',    '145',                      RED),
    ('Validation Pass Rate',f'{val_pass_rate:.0f}%',    GREEN),
]
 
for i, (label, value, color) in enumerate(metrics):
    col = i + 1
    # Label box
    lbl_cell = ws1.cell(row=5, column=col, value=label)
    lbl_cell.font      = Font(name=FONT_NAME, size=9, bold=True, color=WHITE)
    lbl_cell.fill      = fill(MID_BLUE)
    lbl_cell.alignment = center()
    lbl_cell.border    = THIN_BORDER
    ws1.row_dimensions[5].height = 18
    # Value box
    val_cell = ws1.cell(row=6, column=col, value=value)
    val_cell.font      = Font(name=FONT_NAME, size=18, bold=True, color=color)
    val_cell.fill      = fill(LIGHT_BLUE)
    val_cell.alignment = center()
    val_cell.border    = THIN_BORDER
    ws1.row_dimensions[6].height = 36
 
ws1.row_dimensions[7].height = 12
 
# Findings table 
ws1.merge_cells('A8:G8')
sec = ws1['A8']
sec.value     = 'Key Findings Summary'
sec.font      = Font(name=FONT_NAME, size=12, bold=True, color=DARK_BLUE)
sec.alignment = left()
ws1.row_dimensions[8].height = 22
 
apply_header_row(ws1, 9,
    ['Notebook', 'Finding', 'Records Affected', 'Severity', 'Action Taken'],
    col_widths=[22, 52, 20, 14, 36])
 
findings = [
    ('NB1 — Temporal',  '1,997 orders have ship date before purchase date',         '1,997 (9.1%)', 'Medium',   'Flagged IS_ANOMALY=True, retained for revenue analysis'),
    ('NB1 — Temporal',  'Social media channel highest anomaly rate (12.7%)',         '323 orders',   'Info',     'Documented — likely pre-order driven'),
    ('NB1 — Temporal',  'India, Switzerland, South Korea above average anomaly rate','669 orders',   'Info',     'Documented for regional analysis context'),
    ('NB2 — Price',     '$0 price orders — all on website, one product ID cluster',  '29 (0.13%)',   'Medium',   'Flagged IS_ZERO_PRICE=True, excluded from revenue'),
    ('NB2 — Price',     'PS5 orders from GB likely GBP recorded as USD',             '957 outliers', 'High',     'Flagged IS_CURRENCY_SUSPECT=True for manual review'),
    ('NB2 — Price',     '1,430 orders outside per-product IQR price fence',          '1,430 (6.5%)', 'Medium',   'Flagged IS_PRICE_OUTLIER=True'),
    ('NB3 — Entity',    'Duplicate product name: 27inches vs 27in 4K monitor',       '61 orders',    'Medium',   'Standardised to canonical name'),
    ('NB3 — Entity',    '46 product IDs reduced to 9 canonical IDs',                 'All records',  'Medium',   'ID mapping table created and applied'),
    ('NB4 — Validation','US (10,294 orders) missing from region lookup table',        '10,294 (47%)', 'Critical', 'All 15 missing countries manually mapped'),
    ('NB4 — Validation','145 duplicate ORDER_IDs — all in January 2020',             '290 records',  'Critical', '145 duplicate rows removed, dataset: 21,719'),
]
 
severity_colors = {
    'Critical': (LIGHT_RED,   RED),
    'High':     (LIGHT_RED,   RED),
    'Medium':   (LIGHT_AMBER, AMBER),
    'Info':     (LIGHT_BLUE,  MID_BLUE),
}
 
for i, (nb, finding, affected, severity, action) in enumerate(findings):
    row = 10 + i
    alt = i % 2 == 0
    bg_sev, fg_sev = severity_colors.get(severity, (LIGHT_GREY, DARK_GREY))
    bg = LIGHT_GREY if alt else WHITE
 
    for col, val in enumerate([nb, finding, affected, severity, action], 1):
        cell = ws1.cell(row=row, column=col, value=val)
        cell.font      = body_font(size=9)
        cell.alignment = left()
        cell.border    = THIN_BORDER
        if col == 4:
            cell.fill = fill(bg_sev)
            cell.font = Font(name=FONT_NAME, size=9, bold=True, color=fg_sev)
            cell.alignment = center()
        else:
            cell.fill = fill(bg)
    ws1.row_dimensions[row].height = 28
 
ws1.freeze_panes = 'A9'
 
# Column widths
for col_letter, width in zip('ABCDEFG', [22, 52, 20, 14, 36, 14, 14]):
    ws1.column_dimensions[col_letter].width = width


# SHEET 2 — TEMPORAL ANOMALIES
ws2 = wb.create_sheet('02 Temporal Anomalies')
ws2.sheet_view.showGridLines = False
 
ws2.merge_cells('A1:F1')
t = ws2['A1']
t.value     = 'Temporal Anomaly Investigation — Notebook 1'
t.font      = Font(name=FONT_NAME, size=14, bold=True, color=DARK_BLUE)
t.alignment = left()
t.fill      = fill(LIGHT_BLUE)
ws2.row_dimensions[1].height = 30
 
ws2.merge_cells('A2:F2')
s = ws2['A2']
s.value     = 'Orders where ship date precedes purchase date — investigation of cause and recommended handling'
s.font      = Font(name=FONT_NAME, size=9, color='666666', italic=True)
s.alignment = left()
ws2.row_dimensions[2].height = 16
ws2.row_dimensions[3].height = 8
 
# Anomaly rate by product
ws2.cell(row=4, column=1, value='Anomaly Rate by Product').font = Font(
    name=FONT_NAME, size=11, bold=True, color=DARK_BLUE)
ws2.row_dimensions[4].height = 20
 
apply_header_row(ws2, 5,
    ['Product', 'Total Orders', 'Anomalous Orders', 'Anomaly Rate (%)', 'vs Overall Avg'],
    col_widths=[32, 16, 18, 18, 16])
 
product_anom = df_master.copy()
product_anom['IS_ANOMALY'] = product_anom['IS_ANOMALY'].fillna(False)
product_anom['IS_ANOMALY'] = product_anom['IS_ANOMALY'].astype(bool)
product_stats = product_anom.groupby('PRODUCT_NAME').agg(
    total     = ('ORDER_ID', 'count'),
    anomalies = ('IS_ANOMALY', 'sum')
).assign(rate=lambda x: (x['anomalies'] / x['total'] * 100).round(1)
).sort_values('rate', ascending=False).reset_index()
 
overall_rate = product_anom['IS_ANOMALY'].mean() * 100
 
for i, row_data in product_stats.iterrows():
    row = 6 + i # pyright: ignore[reportOperatorIssue]
    alt = i % 2 == 0 # pyright: ignore[reportOperatorIssue]
    vs_avg = row_data['rate'] - overall_rate
    apply_data_row(ws2, row,
        [row_data['PRODUCT_NAME'], row_data['total'],
         row_data['anomalies'], row_data['rate'], f'{vs_avg:+.1f}%'], alt=alt)
    # Colour the rate cell
    rate_cell = ws2.cell(row=row, column=4)
    if row_data['rate'] > overall_rate + 1:
        rate_cell.fill = fill(LIGHT_RED)
        rate_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=RED)
    ws2.row_dimensions[row].height = 18
 
ws2.row_dimensions[6 + len(product_stats) + 1].height = 12
 
# Anomaly rate by marketing channel 
offset = 6 + len(product_stats) + 2
ws2.cell(row=offset, column=1, value='Anomaly Rate by Marketing Channel').font = Font(
    name=FONT_NAME, size=11, bold=True, color=DARK_BLUE)
ws2.row_dimensions[offset].height = 20
 
apply_header_row(ws2, offset+1,
    ['Marketing Channel', 'Total Orders', 'Anomalous Orders', 'Anomaly Rate (%)'],
    col_widths=[24, 16, 18, 18])
 
channel_stats = product_anom.groupby('MARKETING_CHANNEL').agg(
    total     = ('ORDER_ID', 'count'),
    anomalies = ('IS_ANOMALY', 'sum')
).assign(rate=lambda x: (x['anomalies'] / x['total'] * 100).round(1)
).sort_values('rate', ascending=False).reset_index()
 
for i, row_data in channel_stats.iterrows():
    row = offset + 2 + i # pyright: ignore[reportOperatorIssue]
    alt = i % 2 == 0 # pyright: ignore[reportOperatorIssue]
    apply_data_row(ws2, row,
        [row_data['MARKETING_CHANNEL'], row_data['total'],
         row_data['anomalies'], row_data['rate']], alt=alt)
    if row_data['rate'] > overall_rate + 1:
        c = ws2.cell(row=row, column=4)
        c.fill = fill(LIGHT_RED)
        c.font = Font(name=FONT_NAME, size=10, bold=True, color=RED)
    ws2.row_dimensions[row].height = 18
 
ws2.freeze_panes = 'A5'
 

# SHEET 3 — PRICE FORENSICS
ws3 = wb.create_sheet('03 Price Forensics')
ws3.sheet_view.showGridLines = False
 
ws3.merge_cells('A1:G1')
t = ws3['A1']
t.value     = 'Price Forensics — Notebook 2'
t.font      = Font(name=FONT_NAME, size=14, bold=True, color=DARK_BLUE)
t.alignment = left()
t.fill      = fill(LIGHT_BLUE)
ws3.row_dimensions[1].height = 30
 
ws3.merge_cells('A2:G2')
s = ws3['A2']
s.value     = 'Price distribution analysis, outlier investigation, and flagging decisions'
s.font      = Font(name=FONT_NAME, size=9, color='666666', italic=True)
s.alignment = left()
ws3.row_dimensions[2].height = 16
ws3.row_dimensions[3].height = 8
 
# Overall price statistics 
ws3.cell(row=4, column=1, value='Overall Price Statistics').font = Font(
    name=FONT_NAME, size=11, bold=True, color=DARK_BLUE)
ws3.row_dimensions[4].height = 20
 
apply_header_row(ws3, 5,
    ['Metric', 'Value'],
    col_widths=[30, 20])
 
price_stats = [
    ('Total orders with price data', f'{df_master["USD_PRICE"].count():,}'),
    ('Mean price',                   f'${df_master["USD_PRICE"].mean():.2f}'),
    ('Median price',                 f'${df_master["USD_PRICE"].median():.2f}'),
    ('Std deviation',                f'${df_master["USD_PRICE"].std():.2f}'),
    ('Min price',                    f'${df_master["USD_PRICE"].min():.2f}'),
    ('Max price',                    f'${df_master["USD_PRICE"].max():.2f}'),
    ('$0 price orders',              f'{int(df_master["IS_ZERO_PRICE"].sum()) if "IS_ZERO_PRICE" in df_master.columns else 29}'),
    ('Price outliers (> IQR fence)', f'{int(df_master["IS_PRICE_OUTLIER"].sum()) if "IS_PRICE_OUTLIER" in df_master.columns else 1430}'),
]
 
for i, (metric, value) in enumerate(price_stats):
    row = 6 + i
    apply_data_row(ws3, row, [metric, value], alt=i % 2 == 0)
    ws3.row_dimensions[row].height = 18
 
# Per-product price stats
offset3 = 6 + len(price_stats) + 2
ws3.cell(row=offset3, column=1, value='Price Statistics by Product').font = Font(
    name=FONT_NAME, size=11, bold=True, color=DARK_BLUE)
ws3.row_dimensions[offset3].height = 20
 
apply_header_row(ws3, offset3+1,
    ['Product', 'Orders', 'Median ($)', 'Mean ($)', 'Min ($)', 'Max ($)', 'Outlier Rate (%)'],
    col_widths=[32, 10, 13, 13, 13, 13, 17])
 
prod_price = df_master.groupby('PRODUCT_NAME')['USD_PRICE'].agg(
    orders=('count'),
    median=('median'),
    mean=('mean'),
    min=('min'),
    max=('max')
).reset_index().sort_values('median', ascending=False)
 
# Merge outlier rate if column exists
if 'IS_PRICE_OUTLIER' in df_master.columns:
    outlier_rates = df_master.groupby('PRODUCT_NAME')['IS_PRICE_OUTLIER'].mean() * 100
    prod_price = prod_price.merge(outlier_rates.rename('outlier_rate'),
                                   on='PRODUCT_NAME', how='left')
else:
    prod_price['outlier_rate'] = 0
 
for i, row_data in enumerate(prod_price.iterrows()):
    row = offset3 + 2 + i
    alt = i % 2 == 0
    _, row_data = row_data
    apply_data_row(ws3, row, [
        row_data['PRODUCT_NAME'],
        int(row_data['orders']),
        f"${row_data['median']:.2f}",
        f"${row_data['mean']:.2f}",
        f"${row_data['min']:.2f}",
        f"${row_data['max']:.2f}",
        f"{row_data['outlier_rate']:.1f}%"
    ], alt=alt)
    # Highlight high outlier rates
    if row_data['outlier_rate'] > 10:
        c = ws3.cell(row=row, column=7)
        c.fill = fill(LIGHT_RED)
        c.font = Font(name=FONT_NAME, size=10, bold=True, color=RED)
    ws3.row_dimensions[row].height = 18
 
ws3.freeze_panes = 'A5'
 

# SHEET 4 — ENTITY RESOLUTION
ws4 = wb.create_sheet('04 Entity Resolution')
ws4.sheet_view.showGridLines = False
 
ws4.merge_cells('A1:E1')
t = ws4['A1']
t.value     = 'Entity Resolution — Notebook 3'
t.font      = Font(name=FONT_NAME, size=14, bold=True, color=DARK_BLUE)
t.alignment = left()
t.fill      = fill(LIGHT_BLUE)
ws4.row_dimensions[1].height = 30
 
ws4.merge_cells('A2:E2')
s = ws4['A2']
s.value     = 'Product name standardisation and product ID consolidation mapping tables'
s.font      = Font(name=FONT_NAME, size=9, color='666666', italic=True)
s.alignment = left()
ws4.row_dimensions[2].height = 16
ws4.row_dimensions[3].height = 8
 
# Product name mapping
ws4.cell(row=4, column=1, value='Product Name Mapping').font = Font(
    name=FONT_NAME, size=11, bold=True, color=DARK_BLUE)
ws4.row_dimensions[4].height = 20
 
apply_header_row(ws4, 5,
    ['Original Name', 'Canonical Name', 'Is Duplicate?'],
    col_widths=[38, 38, 16])
 
for i, row_data in df_name_map.iterrows():
    row = 6 + i # pyright: ignore[reportOperatorIssue]
    alt = i % 2 == 0 # pyright: ignore[reportOperatorIssue]
    apply_data_row(ws4, row, [
        row_data['original_name'],
        row_data['canonical_name'],
        'YES' if row_data['is_duplicate'] else 'No'
    ], alt=alt)
    if row_data['is_duplicate']:
        for col in range(1, 4):
            c = ws4.cell(row=row, column=col)
            c.fill = fill(LIGHT_AMBER)
        ws4.cell(row=row, column=3).font = Font(
            name=FONT_NAME, size=10, bold=True, color=AMBER)
    ws4.row_dimensions[row].height = 18
 
# Product ID mapping 
offset4 = 6 + len(df_name_map) + 2
ws4.cell(row=offset4, column=1, value='Product ID Mapping (summary — canonical ID per product)').font = Font(
    name=FONT_NAME, size=11, bold=True, color=DARK_BLUE)
ws4.row_dimensions[offset4].height = 20
 
apply_header_row(ws4, offset4+1,
    ['Product', 'Canonical ID', 'Total ID Variants', 'Total Orders'],
    col_widths=[38, 16, 18, 14])
 
id_summary = df_id_map.groupby('product_name_clean').agg(
    canonical_id   = ('canonical_product_id', 'first'),
    total_variants = ('original_product_id', 'count'),
    total_orders   = ('order_count', 'sum')
).reset_index().sort_values('total_orders', ascending=False)
 
for i, (_, row_data) in enumerate(id_summary.iterrows()):
    row = offset4 + 2 + i
    alt = (offset4 + 2 + i) % 2 == 0
    apply_data_row(ws4, row, [
        row_data['product_name_clean'],
        row_data['canonical_id'],
        int(row_data['total_variants']),
        f"{int(row_data['total_orders']):,}"
    ], alt=alt)
    if row_data['total_variants'] > 3:
        c = ws4.cell(row=row, column=3)
        c.fill = fill(LIGHT_AMBER)
        c.font = Font(name=FONT_NAME, size=10, bold=True, color=AMBER)
    ws4.row_dimensions[row].height = 18
 
ws4.freeze_panes = 'A5'
 

# SHEET 5 — VALIDATION RESULTS
ws5 = wb.create_sheet('05 Validation Results')
ws5.sheet_view.showGridLines = False
 
ws5.merge_cells('A1:E1')
t = ws5['A1']
t.value     = 'Great Expectations Validation Results — Notebook 4'
t.font      = Font(name=FONT_NAME, size=14, bold=True, color=DARK_BLUE)
t.alignment = left()
t.fill      = fill(LIGHT_BLUE)
ws5.row_dimensions[1].height = 30
 
# Summary metrics
total_exp = len(df_validation)
passed    = int(df_validation['passed'].sum())
failed    = total_exp - passed
pass_rate = passed / total_exp * 100
 
ws5.merge_cells('A2:E2')
s = ws5['A2']
s.value     = f'Total: {total_exp} expectations   |   Passed: {passed}   |   Failed: {failed}   |   Pass rate: {pass_rate:.0f}%'
s.font      = Font(name=FONT_NAME, size=10, bold=True,
                   color=GREEN if failed == 0 else AMBER)
s.alignment = center()
s.fill      = fill(LIGHT_GREEN if failed == 0 else LIGHT_AMBER)
ws5.row_dimensions[2].height = 22
ws5.row_dimensions[3].height = 8
 
apply_header_row(ws5, 4,
    ['Category', 'Expectation', 'Result', 'Run Date', 'Dataset'],
    col_widths=[20, 52, 12, 14, 30])
 
for i, row_data in df_validation.iterrows():
    row = 5 + i # pyright: ignore[reportOperatorIssue]
    alt = i % 2 == 0 # pyright: ignore[reportOperatorIssue]
    passed_val = bool(row_data['passed'])
    apply_data_row(ws5, row, [
        row_data['category'],
        row_data['expectation'],
        '✅ PASS' if passed_val else '❌ FAIL',
        row_data.get('run_date', ''),
        row_data.get('dataset', '')
    ], alt=alt)
    result_cell = ws5.cell(row=row, column=3)
    if passed_val:
        result_cell.fill = fill(LIGHT_GREEN)
        result_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=GREEN)
    else:
        result_cell.fill = fill(LIGHT_RED)
        result_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=RED)
    result_cell.alignment = center()
    ws5.row_dimensions[row].height = 18
 
ws5.freeze_panes = 'A4'
 
# Final column widths for ws5 
for col_letter, width in zip('ABCDE', [20, 52, 12, 14, 30]):
    ws5.column_dimensions[col_letter].width = width
 
# SAVE

wb.save(OUTPUT_PATH)
print(f'\n✅ Report saved → {OUTPUT_PATH}')
print(f'   Sheets: {wb.sheetnames}')