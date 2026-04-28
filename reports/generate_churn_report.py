"""
generate_churn_report.py
-------------------------
Generates the Project 04 Churn and Survival Analysis Excel report.
Run from the project root:
    python reports/generate_churn_report.py
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT        = Path(__file__).resolve().parent.parent
PROCESSED   = ROOT / 'data' / 'processed'
OUTPUT_PATH = ROOT / 'reports' / 'excel_outputs' / 'churn_report.xlsx'
OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

print('Loading processed data...')
survival_df  = pd.read_csv(PROCESSED / 'survival_data.csv')
km_summary   = pd.read_csv(PROCESSED / 'km_survival_summary.csv')
cox_hr       = pd.read_csv(PROCESSED / 'cox_hazard_ratios.csv')
risk_scores  = pd.read_csv(PROCESSED / 'churn_risk_scores.csv')
priority_sum = pd.read_csv(PROCESSED / 'retention_priority_summary.csv', index_col=0)
config       = pd.read_csv(PROCESSED / 'churn_config.csv').iloc[0]

print(f'  Survival data:    {len(survival_df):,} customers')
print(f'  KM summary:       {len(km_summary)} groups')
print(f'  Cox hazard ratios:{len(cox_hr)} features')
print(f'  Risk scores:      {len(risk_scores):,} customers')

DARK_BLUE   = '1F4E79'
MID_BLUE    = '2E75B6'
LIGHT_BLUE  = 'D6E4F0'
WHITE       = 'FFFFFF'
DARK_GREY   = '404040'
LIGHT_GREY  = 'F2F2F2'
GREEN       = '1E6B3C'
LIGHT_GREEN = 'E2EFDA'
RED         = 'C00000'
LIGHT_RED   = 'FCE4D6'
AMBER       = 'ED7D31'
LIGHT_AMBER = 'FFF2CC'
FONT        = 'Arial'

thin        = Side(style='thin', color='CCCCCC')
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(size=10, bold=True, color=WHITE):
    return Font(name=FONT, size=size, bold=bold, color=color)

def body(size=10, bold=False, color=DARK_GREY):
    return Font(name=FONT, size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def write_header_row(ws, row, headers, col_widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = hdr(10)
        c.fill = fill(MID_BLUE)
        c.alignment = center()
        c.border = THIN_BORDER
        if col_widths and i <= len(col_widths):
            ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]

def write_data_row(ws, row, values, alt=False, bold=False):
    bg = LIGHT_GREY if alt else WHITE
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = body(bold=bold)
        c.fill = fill(bg)
        c.alignment = left()
        c.border = THIN_BORDER

def title_block(ws, row, text, subtitle=None, ncols=7):
    ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
    c = ws[f'A{row}']
    c.value = text
    c.font = Font(name=FONT, size=14, bold=True, color=DARK_BLUE)
    c.fill = fill(LIGHT_BLUE)
    c.alignment = left()
    ws.row_dimensions[row].height = 30
    if subtitle:
        ws.merge_cells(f'A{row+1}:{get_column_letter(ncols)}{row+1}')
        s = ws[f'A{row+1}']
        s.value = subtitle
        s.font = Font(name=FONT, size=9, color='666666', italic=True)
        s.alignment = left()
        ws.row_dimensions[row+1].height = 16

def section_label(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=11, bold=True, color=DARK_BLUE)
    ws.row_dimensions[row].height = 22

wb = Workbook()

# Sheet 1 - Executive Summary
ws1 = wb.active
ws1.title = '01 Executive Summary'
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 6

for col in range(1, 8):
    ws1.cell(row=2, column=col).fill = fill(DARK_BLUE)
ws1.merge_cells('A2:G2')
t = ws1['A2']
t.value = 'GameZone Analytics - Project 04 Churn and Survival Analysis Report'
t.font = Font(name=FONT, size=16, bold=True, color=WHITE)
t.alignment = center()
ws1.row_dimensions[2].height = 36

ws1.merge_cells('A3:G3')
s = ws1['A3']
churn_rate = float(config['churn_rate_pct'])
n_customers = int(config['n_customers'])
s.value = (f'Generated: {datetime.now().strftime("%d %B %Y")}   |   '
           f'{n_customers:,} customers analysed   |   '
           f'Churn threshold: {int(config["churn_threshold_days"])} days')
s.font = Font(name=FONT, size=9, color='888888')
s.alignment = center()
s.fill = fill(LIGHT_GREY)
ws1.row_dimensions[3].height = 18
ws1.row_dimensions[4].height = 10

n_churned  = int(config['n_churned'])
n_active   = n_customers - n_churned
p1_revenue = priority_sum.loc['High CLV / High Risk', 'total_revenue'] \
    if 'High CLV / High Risk' in priority_sum.index else 0
p1_customers = int(priority_sum.loc['High CLV / High Risk', 'customers']) \
    if 'High CLV / High Risk' in priority_sum.index else 0
concordance = 0.89

metrics_display = [
    ('Total Customers',     f'{n_customers:,}',            DARK_BLUE),
    ('Churned',             f'{n_churned:,} ({churn_rate:.1f}%)', RED),
    ('Active/Censored',     f'{n_active:,}',               GREEN),
    ('Cox Concordance',     f'{concordance}',              GREEN),
    ('P1 Customers',        f'{p1_customers:,}',           AMBER),
    ('P1 Revenue at Stake', f'${p1_revenue:,.0f}',         RED),
    ('Churn Threshold',     f'{int(config["churn_threshold_days"])} days', DARK_BLUE),
]

for i, (label, value, color) in enumerate(metrics_display):
    col = i + 1
    lbl = ws1.cell(row=5, column=col, value=label)
    lbl.font = Font(name=FONT, size=8, bold=True, color=WHITE)
    lbl.fill = fill(MID_BLUE)
    lbl.alignment = center()
    lbl.border = THIN_BORDER
    ws1.row_dimensions[5].height = 18

    val = ws1.cell(row=6, column=col, value=value)
    val.font = Font(name=FONT, size=12, bold=True, color=color)
    val.fill = fill(LIGHT_BLUE)
    val.alignment = center()
    val.border = THIN_BORDER
    ws1.row_dimensions[6].height = 32

ws1.row_dimensions[7].height = 10
section_label(ws1, 8, 'Key Findings Summary')
write_header_row(ws1, 9,
    ['Notebook', 'Finding', 'Key Metric', 'Implication'],
    col_widths=[22, 52, 22, 36])

findings = [
    ('NB1 - Churn Labelling',
     'Churn threshold: 180 days (90th pct of inter-purchase gaps)',
     '180-day threshold',
     'Data-driven threshold accounts for gaming hardware repurchase cycles'),
    ('NB1 - Churn Labelling',
     '72.5% churn rate reflects product category, not business failure',
     '72.5% churn',
     'Gaming hardware buyers have 2-4 year repurchase cycles by nature'),
    ('NB1 - Churn Labelling',
     'Lapsed RFM segment = 100% churn - validates both methodologies',
     '100% Lapsed churn',
     'RFM segmentation and survival analysis identify identical customers'),
    ('NB1 - Churn Labelling',
     'Email channel has lowest churn rate (66.5%) despite lowest CLV',
     'Email: 66.5% churn',
     'Email attracts lower-spend but longer-tenure customers'),
    ('NB2 - Kaplan-Meier',
     'Median survival 1 day - 91% of customers are one-time buyers',
     'Median: 1 day',
     'Survival signal concentrated in Champions segment only'),
    ('NB2 - Kaplan-Meier',
     'Champions median survival 28 days - only segment with signal',
     'Champions: 28 days',
     'All other segments churn on day 1 - no differential survival'),
    ('NB2 - Kaplan-Meier',
     'All three stratifications are statistically significant',
     'All p < 0.05',
     'Segment, channel, and region all independently predict survival time'),
    ('NB3 - Cox Model',
     'Concordance 0.89 - strong model accuracy despite dataset constraints',
     'Concordance: 0.89',
     'Model correctly ranks churn timing in 89% of customer pairs'),
    ('NB3 - Cox Model',
     'RFM_score HR 0.524 - strongest protective factor against churn',
     'HR = 0.524',
     'Each SD increase in RFM score cuts churn risk by nearly half'),
    ('NB3 - Cox Model',
     'High-ticket buyers churn 26% slower (HR 0.744)',
     'HR = 0.744',
     'PS5/laptop buyers are more engaged - not just one-off purchasers'),
    ('NB3 - Cox Model',
     'Higher avg order value increases churn risk (HR 1.297)',
     'HR = 1.297',
     'Likely reflects currency-suspect GB orders from Project 01'),
    ('NB4 - Risk Scoring',
     '3,210 High CLV / High Risk customers - $1.35M revenue at stake',
     '$1,350,703 at risk',
     'Priority 1 retention target - immediate personalised outreach'),
]

nb_colors = {
    'NB1 - Churn Labelling': (LIGHT_BLUE,  MID_BLUE),
    'NB2 - Kaplan-Meier':    (LIGHT_AMBER, AMBER),
    'NB3 - Cox Model':       (LIGHT_GREEN, GREEN),
    'NB4 - Risk Scoring':    (LIGHT_RED,   RED),
}

for i, (nb, finding, metric, impl) in enumerate(findings):
    row = 10 + i
    alt = i % 2 == 0
    bg, fg = nb_colors.get(nb, (LIGHT_GREY, DARK_GREY))
    write_data_row(ws1, row, [nb, finding, metric, impl], alt=alt)
    nb_cell = ws1.cell(row=row, column=1)
    nb_cell.fill = fill(bg)
    nb_cell.font = Font(name=FONT, size=9, bold=True, color=fg)
    nb_cell.alignment = center()
    ws1.row_dimensions[row].height = 26

ws1.freeze_panes = 'A9'
for col_letter, width in zip('ABCDEFG', [22, 52, 22, 36, 12, 12, 12]):
    ws1.column_dimensions[col_letter].width = width

# Sheet 2 - Kaplan-Meier Summary
ws2 = wb.create_sheet('02 Kaplan-Meier Summary')
ws2.sheet_view.showGridLines = False

title_block(ws2, 1, 'Kaplan-Meier Survival Analysis - Notebook 2',
    'Median survival times and log-rank test results by group')

section_label(ws2, 4, 'Survival summary by group')
write_header_row(ws2, 5,
    ['Group Type', 'Group', 'N', 'Events', 'Median Survival (days)',
     'Log-rank p-value', 'Significant?'],
    col_widths=[18, 22, 10, 10, 22, 18, 14])

group_type_colors = {
    'Overall': (LIGHT_BLUE,  MID_BLUE),
    'Segment': (LIGHT_AMBER, AMBER),
    'Channel': (LIGHT_GREEN, GREEN),
    'Region':  (LIGHT_GREY,  DARK_GREY),
}

for i, row_data in km_summary.iterrows():
    row = 6 + i
    alt = i % 2 == 0
    p   = row_data.get('logrank_p')
    sig = 'Yes' if pd.notna(p) and p < 0.05 else ('No' if pd.notna(p) else 'N/A')
    p_display = f'{p:.6f}' if pd.notna(p) else 'N/A'
    median = row_data['median_survival_days']

    write_data_row(ws2, row, [
        row_data['group_type'],
        row_data['group'],
        f"{int(row_data['n']):,}",
        f"{int(row_data['events']):,}",
        f'{median:.0f}',
        p_display,
        sig
    ], alt=alt)

    bg, fg = group_type_colors.get(row_data['group_type'], (WHITE, DARK_GREY))
    ws2.cell(row=row, column=1).fill = fill(bg)
    ws2.cell(row=row, column=1).font = Font(name=FONT, size=10, bold=True, color=fg)

    if sig == 'Yes':
        c = ws2.cell(row=row, column=7)
        c.fill = fill(LIGHT_GREEN)
        c.font = Font(name=FONT, size=10, bold=True, color=GREEN)
        c.alignment = center()

    ws2.row_dimensions[row].height = 18

ws2.freeze_panes = 'A5'

# Sheet 3 - Cox Hazard Ratios
ws3 = wb.create_sheet('03 Cox Hazard Ratios')
ws3.sheet_view.showGridLines = False

title_block(ws3, 1, 'Cox Proportional Hazards Model - Notebook 3',
    'Concordance: 0.89 | Penalizer: 0.1 | 13,778 customers | 10 features')

section_label(ws3, 4, 'Hazard ratios - all features')
write_header_row(ws3, 5,
    ['Feature', 'Hazard Ratio', 'HR Lower 95%', 'HR Upper 95%',
     'p-value', 'Significant?', 'Direction', 'Business Interpretation'],
    col_widths=[22, 14, 14, 14, 10, 14, 14, 48])

interpretations = {
    'RFM_score':          'Each SD increase in RFM halves churn risk - strongest protective factor',
    'is_high_ticket_buyer': 'PS5/laptop buyers churn 26% slower - more engaged customers',
    'is_email':           'Email channel customers churn 12% slower than reference group',
    'is_apac':            'APAC customers churn 9% slower - strong engagement in region',
    'is_emea':            'EMEA customers churn 7% slower than Americas reference group',
    'avg_order_value':    'Higher prices linked to 30% faster churn - likely currency anomaly',
    'total_orders':       '5% higher churn per additional order - statistical observation period effect',
    'is_affiliate':       'No significant difference from reference channel (p=0.42)',
    'is_direct':          'No significant difference from reference channel (p=0.71)',
    'is_latam':           'Directionally protective but not significant at 5% level (p=0.07)',
}

cox_sorted = cox_hr.sort_values('hazard_ratio', ascending=False)

for i, row_data in cox_sorted.iterrows():
    row = 6 + list(cox_sorted.index).index(i)
    alt = row % 2 == 0
    hr  = row_data['hazard_ratio']
    sig = 'Yes' if row_data['significant'] else 'No'
    direction = 'Increases risk' if hr > 1 else 'Protective'
    feat = row_data['feature']

    write_data_row(ws3, row, [
        feat,
        f'{hr:.3f}x',
        f"{row_data['hr_lower_95']:.3f}",
        f"{row_data['hr_upper_95']:.3f}",
        f"{row_data['p']:.4f}" if row_data['p'] > 0 else '<0.0001',
        sig,
        direction,
        interpretations.get(feat, '')
    ], alt=alt)

    hr_cell = ws3.cell(row=row, column=2)
    dir_cell = ws3.cell(row=row, column=7)
    if hr > 1:
        hr_cell.fill = fill(LIGHT_RED)
        hr_cell.font = Font(name=FONT, size=10, bold=True, color=RED)
        dir_cell.fill = fill(LIGHT_RED)
        dir_cell.font = Font(name=FONT, size=10, color=RED)
    else:
        hr_cell.fill = fill(LIGHT_GREEN)
        hr_cell.font = Font(name=FONT, size=10, bold=True, color=GREEN)
        dir_cell.fill = fill(LIGHT_GREEN)
        dir_cell.font = Font(name=FONT, size=10, color=GREEN)

    ws3.row_dimensions[row].height = 22

ws3.freeze_panes = 'A5'

# Sheet 4 - Retention Priority Matrix
ws4 = wb.create_sheet('04 Retention Priority')
ws4.sheet_view.showGridLines = False

title_block(ws4, 1, 'Retention Priority Matrix - Notebook 4',
    'Customers segmented by CLV tier and churn risk band')

section_label(ws4, 4, 'Priority quadrant summary')
write_header_row(ws4, 5,
    ['Priority', 'Quadrant', 'Customers', 'Avg Spend ($)',
     'Avg Risk Score', 'Revenue at Stake ($)', 'Recommended Action'],
    col_widths=[12, 26, 14, 16, 16, 20, 52])

priority_config = [
    ('P1', 'High CLV / High Risk',
     'PRIORITY 1 - Immediate personalised retention outreach. Highest ROI target.'),
    ('P2', 'High CLV / Low Risk',
     'PRIORITY 2 - Monitor quarterly. Valuable and stable - nurture the relationship.'),
    ('P3', 'Low CLV / High Risk',
     'PRIORITY 3 - Single low-cost re-engagement email. Do not over-invest.'),
    ('P4', 'Low CLV / Low Risk',
     'PRIORITY 4 - No active retention investment required.'),
]

priority_colors = {
    'P1': (LIGHT_RED,   RED),
    'P2': (LIGHT_AMBER, AMBER),
    'P3': (LIGHT_GREY,  DARK_GREY),
    'P4': (LIGHT_GREY,  DARK_GREY),
}

for i, (priority, quadrant, action) in enumerate(priority_config):
    row = 6 + i
    bg, fg = priority_colors.get(priority, (WHITE, DARK_GREY))
    if quadrant in priority_sum.index:
        r = priority_sum.loc[quadrant]
        write_data_row(ws4, row, [
            priority,
            quadrant,
            f"{int(r['customers']):,}",
            f"${r['avg_spend']:,.2f}",
            f"{r['avg_risk_score']:.1f}",
            f"${r['total_revenue']:,.0f}",
            action
        ])
        for col in range(1, 8):
            c = ws4.cell(row=row, column=col)
            c.fill = fill(bg)
            c.font = Font(name=FONT, size=10,
                         bold=(col <= 2), color=fg)
        ws4.row_dimensions[row].height = 28

ws4.row_dimensions[11].height = 10
section_label(ws4, 12, 'Risk tier distribution')
write_header_row(ws4, 13,
    ['Risk Tier', 'Score Range', 'Customers', 'Avg Spend ($)',
     'Churn Rate', 'Revenue at Risk ($)'],
    col_widths=[14, 14, 14, 16, 14, 20])

tier_data = risk_scores.groupby('risk_tier', observed=True).agg(
    customers     = ('USER_ID',          'count'),
    avg_spend     = ('total_spend',       'mean'),
    churn_rate    = ('churned',           'mean'),
    total_revenue = ('total_spend',       'sum')
).round(2)

tier_ranges   = {'Low': '0-25', 'Medium': '26-50', 'High': '51-75', 'Critical': '76-100'}
tier_bg_colors = {
    'Low':      (LIGHT_GREEN, GREEN),
    'Medium':   (LIGHT_BLUE,  MID_BLUE),
    'High':     (LIGHT_AMBER, AMBER),
    'Critical': (LIGHT_RED,   RED),
}

for i, tier in enumerate(['Low', 'Medium', 'High', 'Critical']):
    if tier not in tier_data.index:
        continue
    row = 14 + i
    r   = tier_data.loc[tier]
    bg, fg = tier_bg_colors[tier]
    write_data_row(ws4, row, [
        tier,
        tier_ranges.get(tier, ''),
        f"{int(r['customers']):,}",
        f"${r['avg_spend']:,.2f}",
        f"{r['churn_rate']:.1%}",
        f"${r['total_revenue']:,.0f}"
    ])
    for col in range(1, 7):
        c = ws4.cell(row=row, column=col)
        c.fill = fill(bg)
        c.font = Font(name=FONT, size=10, bold=(col == 1), color=fg)
    ws4.row_dimensions[row].height = 20

ws4.freeze_panes = 'A5'

# Sheet 5 - Churn Config and Methodology
ws5 = wb.create_sheet('05 Methodology')
ws5.sheet_view.showGridLines = False

title_block(ws5, 1, 'Methodology and Configuration', ncols=5)

section_label(ws5, 4, 'Churn definition parameters')
write_header_row(ws5, 5, ['Parameter', 'Value'], col_widths=[36, 36])

config_rows = [
    ('Churn threshold (days)',         str(int(config['churn_threshold_days']))),
    ('Threshold basis',                '90th percentile of inter-purchase gaps among repeat buyers'),
    ('Minimum floor applied',          '180 days - accounts for gaming hardware repurchase cycles'),
    ('Observation window start',       str(config['obs_start'])),
    ('Observation window end',         str(config['obs_end'])),
    ('Total observation days',         str(int(config['obs_days']))),
    ('Total customers',                f"{int(config['n_customers']):,}"),
    ('Churned customers',              f"{int(config['n_churned']):,} ({float(config['churn_rate_pct']):.1f}%)"),
    ('Active/Censored customers',      f"{int(config['n_customers']) - int(config['n_churned']):,}"),
]

for i, (param, value) in enumerate(config_rows):
    row = 6 + i
    write_data_row(ws5, row, [param, value], alt=i % 2 == 0)
    ws5.row_dimensions[row].height = 20

offset5 = 6 + len(config_rows) + 2
section_label(ws5, offset5, 'Survival analysis methodology')
write_header_row(ws5, offset5+1,
    ['Method', 'Purpose', 'Key Output'],
    col_widths=[24, 44, 36])

methods = [
    ('Kaplan-Meier estimator',
     'Non-parametric survival curves - no distribution assumptions',
     'Survival probability at each time point per group'),
    ('Log-rank test',
     'Tests whether survival curves differ significantly between groups',
     'p-value confirming statistical significance of group differences'),
    ('Cox proportional hazards',
     'Multivariate regression controlling all features simultaneously',
     'Hazard ratio per feature - independent effect on churn timing'),
    ('Schoenfeld residuals test',
     'Validates proportional hazards assumption for each feature',
     '9 of 10 features pass - total_orders violates assumption'),
    ('Churn risk scoring',
     'Composite score combining Cox hazard, recency, and RFM (40/40/20%)',
     'Risk score 0-100 and priority quadrant per customer'),
]

for i, (method, purpose, output) in enumerate(methods):
    row = offset5 + 2 + i
    write_data_row(ws5, row, [method, purpose, output], alt=i % 2 == 0)
    ws5.row_dimensions[row].height = 24

ws5.freeze_panes = 'A5'

wb.save(OUTPUT_PATH)
print(f'\nReport saved to {OUTPUT_PATH}')
print(f'Sheets: {wb.sheetnames}')
