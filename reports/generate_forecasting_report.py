"""
generate_forecasting_report.py
--------------------------------
Generates the Project 03 Demand Forecasting Excel report.
Run from the project root:
    python reports/generate_forecasting_report.py
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT        = Path(__file__).resolve().parent.parent
PROCESSED   = ROOT / 'data' / 'processed'
OUTPUT_PATH = ROOT / 'reports' / 'excel_outputs' / 'forecasting_report.xlsx'
OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

print('Loading processed data...')
weekly      = pd.read_csv(PROCESSED / 'ts_features_weekly.csv', parse_dates=['week_start'])
prophet_fc  = pd.read_csv(PROCESSED / 'prophet_forecast.csv',  parse_dates=['ds'])
lgbm_fc     = pd.read_csv(PROCESSED / 'lgbm_forecast.csv',     parse_dates=['week_start'])
metrics     = pd.read_csv(PROCESSED / 'model_metrics.csv')
comparison  = pd.read_csv(PROCESSED / 'model_comparison.csv',  parse_dates=['week_start'])

print(f'  Weekly time series: {len(weekly):,} rows')
print(f'  Prophet forecast:   {len(prophet_fc):,} rows')
print(f'  LightGBM forecast:  {len(lgbm_fc):,} rows')
print(f'  Comparison table:   {len(comparison):,} rows')

# Style constants
DARK_BLUE   = '1F4E79'
MID_BLUE    = '2E75B6'
LIGHT_BLUE  = 'D6E4F0'
WHITE       = 'FFFFFF'
DARK_GREY   = '404040'
LIGHT_GREY  = 'F2F2F2'
GREEN       = '1E6B3C'
LIGHT_GREEN = 'E2EFDA'
RED         = 'C00000'
LIGHT_RED   = 'FCE4D6'
AMBER       = 'ED7D31'
LIGHT_AMBER = 'FFF2CC'
FONT        = 'Arial'

thin        = Side(style='thin',   color='CCCCCC')
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(size=10, bold=True, color=WHITE):
    return Font(name=FONT, size=size, bold=bold, color=color)

def body(size=10, bold=False, color=DARK_GREY):
    return Font(name=FONT, size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def write_header_row(ws, row, headers, col_widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font      = hdr(10)
        c.fill      = fill(MID_BLUE)
        c.alignment = center()
        c.border    = THIN_BORDER
        if col_widths and i <= len(col_widths):
            ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]

def write_data_row(ws, row, values, alt=False, bold=False):
    bg = LIGHT_GREY if alt else WHITE
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font      = body(bold=bold)
        c.fill      = fill(bg)
        c.alignment = left()
        c.border    = THIN_BORDER

def title_block(ws, row, text, subtitle=None, ncols=7):
    ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
    c = ws[f'A{row}']
    c.value     = text
    c.font      = Font(name=FONT, size=14, bold=True, color=DARK_BLUE)
    c.fill      = fill(LIGHT_BLUE)
    c.alignment = left()
    ws.row_dimensions[row].height = 30
    if subtitle:
        ws.merge_cells(f'A{row+1}:{get_column_letter(ncols)}{row+1}')
        s = ws[f'A{row+1}']
        s.value     = subtitle
        s.font      = Font(name=FONT, size=9, color='666666', italic=True)
        s.alignment = left()
        ws.row_dimensions[row+1].height = 16

def section_label(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=11, bold=True, color=DARK_BLUE)
    ws.row_dimensions[row].height = 22

wb = Workbook()

# Sheet 1 - Executive Summary
ws1 = wb.active
ws1.title = '01 Executive Summary'
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 6

for col in range(1, 8):
    ws1.cell(row=2, column=col).fill = fill(DARK_BLUE)
ws1.merge_cells('A2:G2')
t = ws1['A2']
t.value     = 'GameZone Analytics - Project 03 Demand Forecasting Report'
t.font      = Font(name=FONT, size=16, bold=True, color=WHITE)
t.alignment = center()
ws1.row_dimensions[2].height = 36

ws1.merge_cells('A3:G3')
s = ws1['A3']
s.value     = f'Generated: {datetime.now().strftime("%d %B %Y")}   |   112 weekly observations   |   2019-01-07 to 2021-02-22'
s.font      = Font(name=FONT, size=9, color='888888')
s.alignment = center()
s.fill      = fill(LIGHT_GREY)
ws1.row_dimensions[3].height = 18
ws1.row_dimensions[4].height = 10

# Key metric boxes
prophet_row = metrics[metrics['model'] == 'Prophet'].iloc[0]
lgbm_row    = metrics[metrics['model'] == 'LightGBM'].iloc[0]

metrics_display = [
    ('Change-point',      '2020-02-24',              DARK_BLUE),
    ('Weekly observations', f'{len(weekly):,}',      DARK_BLUE),
    ('Prophet train MAPE', f'{prophet_row["train_mape"]:.1f}%', AMBER),
    ('LightGBM CV MAPE',  '11.6%',                  GREEN),
    ('Top SHAP feature',  'roll_std_8',              MID_BLUE),
    ('Recommended model', 'Prophet',                 GREEN),
    ('Dataset verdict',   'Needs richer data',       RED),
]

for i, (label, value, color) in enumerate(metrics_display):
    col = i + 1
    lbl = ws1.cell(row=5, column=col, value=label)
    lbl.font      = Font(name=FONT, size=8, bold=True, color=WHITE)
    lbl.fill      = fill(MID_BLUE)
    lbl.alignment = center()
    lbl.border    = THIN_BORDER
    ws1.row_dimensions[5].height = 18

    val = ws1.cell(row=6, column=col, value=value)
    val.font      = Font(name=FONT, size=13, bold=True, color=color)
    val.fill      = fill(LIGHT_BLUE)
    val.alignment = center()
    val.border    = THIN_BORDER
    ws1.row_dimensions[6].height = 32

ws1.row_dimensions[7].height = 10

section_label(ws1, 8, 'Key Findings Summary')
write_header_row(ws1, 9,
    ['Notebook', 'Finding', 'Key Metric', 'Implication'],
    col_widths=[22, 52, 22, 36])

findings = [
    ('NB1 - Feature Eng',  'PELT detected COVID change-point at 2020-02-24',
     'Revenue: $20K to $60K/wk', 'Train post-COVID regime only for LightGBM'),
    ('NB1 - Feature Eng',  'Lag 1 week has 0.814 correlation with revenue',
     'r = 0.814',               'Recent momentum is the strongest forecast signal'),
    ('NB1 - Feature Eng',  'No weekend effect - revenue flat across all days',
     'Max diff: $302/day',       'is_weekend feature has zero predictive value'),
    ('NB1 - Feature Eng',  'High-ticket % averages 7.3% per week, max 15.8%',
     '7.3% avg / 15.8% max',    'Product mix is a meaningful weekly revenue driver'),
    ('NB2 - Prophet',      'Training MAPE 25.1% on 104 weeks',
     '25.1% train MAPE',        'Reasonable fit; Prophet benchmark established'),
    ('NB2 - Prophet',      'Test MAPE 1,062% due to dataset truncation',
     'Not a valid benchmark',   'Jan-Feb 2021 pre-order exclusion causes collapse'),
    ('NB3 - LightGBM',     '3-fold CV MAPE: 11.6% - honest performance estimate',
     '11.6% CV MAPE',           'Outperforms Prophet on reliable metric'),
    ('NB3 - LightGBM',     'Training MAPE 0.83% signals severe overfitting',
     '0.83% vs 11.6% CV',       '22 features on 40 data points - too few samples'),
    ('NB3 - LightGBM',     'revenue_roll_std_8 is top SHAP feature ($2,730)',
     'SHAP $2,730',              'Revenue volatility predicts future revenue'),
    ('NB3 - LightGBM',     'high_ticket_pct is 2nd most important ($2,633)',
     'SHAP $2,633',              'Product mix confirms Project 02 CLV finding'),
    ('NB4 - Comparison',   'LightGBM outperforms Prophet on shared window',
     '340% vs 710% MAPE',        'Lag features self-correct as actuals fall'),
    ('NB4 - Comparison',   'Dataset insufficient for production forecasting',
     '2-year dataset',           'Recommend 3-5 year dataset rebuild'),
]

nb_colors = {
    'NB1 - Feature Eng': (LIGHT_BLUE,  MID_BLUE),
    'NB2 - Prophet':     (LIGHT_AMBER, AMBER),
    'NB3 - LightGBM':   (LIGHT_GREEN, GREEN),
    'NB4 - Comparison':  (LIGHT_RED,   RED),
}

for i, (nb, finding, metric, impl) in enumerate(findings):
    row = 10 + i
    alt = i % 2 == 0
    bg, fg = nb_colors.get(nb, (LIGHT_GREY, DARK_GREY))
    write_data_row(ws1, row, [nb, finding, metric, impl], alt=alt)
    nb_cell = ws1.cell(row=row, column=1)
    nb_cell.fill = fill(bg)
    nb_cell.font = Font(name=FONT, size=9, bold=True, color=fg)
    nb_cell.alignment = center()
    ws1.row_dimensions[row].height = 26

ws1.freeze_panes = 'A9'
for col_letter, width in zip('ABCDEFG', [22, 52, 22, 36, 12, 12, 12]):
    ws1.column_dimensions[col_letter].width = width

# Sheet 2 - Time Series Data
ws2 = wb.create_sheet('02 Weekly Time Series')
ws2.sheet_view.showGridLines = False

title_block(ws2, 1, 'Weekly Time Series - Feature Engineering Summary',
    'Aggregated from 14,954 orders across 790 daily observations')

section_label(ws2, 4, 'Weekly revenue with key features')
write_header_row(ws2, 5,
    ['Week Start', 'Revenue ($)', 'Order Count', 'Avg Price ($)',
     'High-Ticket %', 'Regime', 'Lag 1 Rev ($)', 'Roll Mean 4 ($)'],
    col_widths=[14, 15, 14, 14, 15, 10, 16, 16])

display_cols = ['week_start', 'revenue', 'order_count', 'avg_price',
                'high_ticket_pct', 'regime', 'revenue_lag_1', 'revenue_roll_mean_4']
available    = [c for c in display_cols if c in weekly.columns]
weekly_disp  = weekly[available].copy()

CHANGE_POINT = pd.to_datetime('2020-02-24')

for i, (_, row_data) in enumerate(weekly_disp.iterrows()):
    row = 6 + i
    alt = i % 2 == 0
    is_cp_week = (pd.to_datetime(row_data['week_start']) >= CHANGE_POINT and
                  pd.to_datetime(row_data['week_start']) < CHANGE_POINT + pd.Timedelta(weeks=2))

    rev = row_data.get('revenue', 0)
    values = [
        str(pd.to_datetime(row_data['week_start']).date()),
        f"${rev:,.0f}" if pd.notna(rev) else '',
        int(row_data['order_count']) if pd.notna(row_data.get('order_count')) else '',
        f"${row_data['avg_price']:.2f}" if pd.notna(row_data.get('avg_price')) else '',
        f"{row_data['high_ticket_pct']:.1%}" if pd.notna(row_data.get('high_ticket_pct')) else '',
        int(row_data['regime']) if pd.notna(row_data.get('regime')) else 0,
        f"${row_data['revenue_lag_1']:,.0f}" if pd.notna(row_data.get('revenue_lag_1')) else '',
        f"${row_data['revenue_roll_mean_4']:,.0f}" if pd.notna(row_data.get('revenue_roll_mean_4')) else '',
    ]
    write_data_row(ws2, row, values, alt=alt)

    if is_cp_week:
        for col in range(1, len(values) + 1):
            c = ws2.cell(row=row, column=col)
            c.fill = fill(LIGHT_RED)
            c.font = Font(name=FONT, size=10, bold=True, color=RED)

    ws2.row_dimensions[row].height = 16

ws2.freeze_panes = 'A5'

# Sheet 3 - Model Performance
ws3 = wb.create_sheet('03 Model Performance')
ws3.sheet_view.showGridLines = False

title_block(ws3, 1, 'Model Performance Comparison',
    'Prophet (Notebook 2) vs LightGBM (Notebook 3)')

section_label(ws3, 4, 'Performance metrics summary')
write_header_row(ws3, 5,
    ['Model', 'Training MAPE', 'CV MAPE', 'Shared Window MAPE',
     'Shared Window MAE', 'Train Weeks', 'Test Weeks', 'Overfitting Risk'],
    col_widths=[14, 16, 14, 22, 20, 14, 12, 18])

# Calculate shared window MAPEs
def calc_mape(actual, predicted):
    mask = np.array(actual) > 0
    a, p = np.array(actual)[mask], np.array(predicted)[mask]
    return np.mean(np.abs((a - p) / a)) * 100

comp = comparison.dropna(subset=['prophet_pred', 'lgbm_pred', 'actual'])
prophet_shared_mape = calc_mape(comp['actual'], comp['prophet_pred'].clip(lower=0))
lgbm_shared_mape    = calc_mape(comp['actual'], comp['lgbm_pred'].clip(lower=0))
prophet_shared_mae  = np.mean(np.abs(comp['actual'] - comp['prophet_pred'].clip(lower=0)))
lgbm_shared_mae     = np.mean(np.abs(comp['actual'] - comp['lgbm_pred'].clip(lower=0)))

model_rows = [
    ('Prophet',  f"{prophet_row['train_mape']:.1f}%", 'Not computed',
     f"{prophet_shared_mape:.1f}%", f"${prophet_shared_mae:,.0f}",
     int(prophet_row['train_weeks']), int(prophet_row['test_weeks']), 'Low'),
    ('LightGBM', f"{lgbm_row['train_mape']:.2f}% (overfit)", '11.6%',
     f"{lgbm_shared_mape:.1f}%", f"${lgbm_shared_mae:,.0f}",
     int(lgbm_row['train_weeks']),  int(lgbm_row['test_weeks']),  'High'),
]

model_colors = {'Prophet': (LIGHT_AMBER, AMBER), 'LightGBM': (LIGHT_GREEN, GREEN)}

for i, row_vals in enumerate(model_rows):
    row = 6 + i
    write_data_row(ws3, row, list(row_vals), alt=False)
    bg, fg = model_colors.get(row_vals[0], (WHITE, DARK_GREY))
    for col in range(1, 9):
        c = ws3.cell(row=row, column=col)
        c.fill = fill(bg)
        c.font = Font(name=FONT, size=10, color=fg, bold=(col == 1))
    ws3.row_dimensions[row].height = 22

ws3.row_dimensions[9].height = 10
section_label(ws3, 10, 'Week-by-week comparison on shared test window')
write_header_row(ws3, 11,
    ['Week', 'Actual ($)', 'Prophet Pred ($)', 'LightGBM Pred ($)',
     'Prophet Error ($)', 'LightGBM Error ($)',
     'Prophet APE (%)', 'LightGBM APE (%)'],
    col_widths=[14, 14, 18, 18, 18, 18, 17, 17])

for i, (_, row_data) in enumerate(comp.iterrows()):
    row = 12 + i
    alt = i % 2 == 0
    p_ape = abs((row_data['actual'] - row_data['prophet_pred']) / row_data['actual'] * 100) if row_data['actual'] > 0 else 0
    l_ape = abs((row_data['actual'] - row_data['lgbm_pred']) / row_data['actual'] * 100) if row_data['actual'] > 0 else 0

    write_data_row(ws3, row, [
        str(row_data['week_start'].date()),
        f"${row_data['actual']:,.0f}",
        f"${row_data['prophet_pred']:,.0f}",
        f"${row_data['lgbm_pred']:,.0f}",
        f"${row_data['actual'] - row_data['prophet_pred']:+,.0f}",
        f"${row_data['actual'] - row_data['lgbm_pred']:+,.0f}",
        f"{p_ape:.1f}%",
        f"{l_ape:.1f}%",
    ], alt=alt)

    better_model = 'lgbm' if l_ape < p_ape else 'prophet'
    better_col   = 8 if better_model == 'lgbm' else 7
    c = ws3.cell(row=row, column=better_col)
    c.fill = fill(LIGHT_GREEN)
    c.font = Font(name=FONT, size=10, bold=True, color=GREEN)
    ws3.row_dimensions[row].height = 18

ws3.freeze_panes = 'A11'

# Sheet 4 - SHAP Feature Importance
ws4 = wb.create_sheet('04 SHAP Feature Importance')
ws4.sheet_view.showGridLines = False

title_block(ws4, 1, 'SHAP Feature Importance - LightGBM Model',
    'Mean absolute SHAP value measures each features average contribution to predictions')

section_label(ws4, 4, 'Feature importance ranking')
write_header_row(ws4, 5,
    ['Rank', 'Feature', 'Mean SHAP Value ($)', 'Category', 'Business Interpretation'],
    col_widths=[8, 28, 22, 18, 52])

shap_data = [
    (1,  'revenue_roll_std_8',    2730, 'Rolling feature',
     'Revenue volatility over 8 weeks - high variance periods predict higher revenue'),
    (2,  'high_ticket_pct',       2633, 'Product mix',
     'Proportion of high-ticket orders - when PS5/laptop orders dominate, revenue spikes'),
    (3,  'month',                 1890, 'Calendar',
     'Monthly seasonality - later months (Q4) predict higher revenue'),
    (4,  'revenue_roll_mean_13',  1713, 'Rolling feature',
     'Quarterly rolling average - captures medium-term revenue momentum'),
    (5,  'order_count_lag_4',     1361, 'Lag feature',
     'Order volume 4 weeks ago - medium-term demand momentum signal'),
    (6,  'revenue_roll_mean_8',   1233, 'Rolling feature',
     'Two-month rolling average - smoothed revenue trend'),
    (7,  'week_of_year',          1211, 'Calendar',
     'Week position in year - captures annual seasonal patterns'),
    (8,  'order_count_lag_2',     1017, 'Lag feature',
     'Order volume 2 weeks ago - short-term demand momentum'),
    (9,  'revenue_lag_8',         1009, 'Lag feature',
     'Revenue 8 weeks ago - longer-horizon momentum signal'),
    (10, 'revenue_lag_1',          894, 'Lag feature',
     'Last weeks revenue - strongest single-period predictor (correlation 0.814)'),
    (11, 'order_count_lag_1',      890, 'Lag feature',
     'Last weeks order count - highly correlated with revenue_lag_1'),
    (12, 'revenue_roll_std_4',     513, 'Rolling feature',
     'Four-week revenue volatility - shorter window volatility signal'),
    (13, 'revenue_lag_4',          492, 'Lag feature',
     'Revenue 4 weeks ago - monthly momentum'),
    (14, 'revenue_roll_mean_4',    396, 'Rolling feature',
     'Four-week rolling average - short-term smoothed trend'),
    (15, 'revenue_lag_2',          382, 'Lag feature',
     'Revenue 2 weeks ago - fortnight momentum'),
    (16, 'month_sin',              266, 'Calendar (cyclical)',
     'Sine encoding of month - treats Dec and Jan as adjacent'),
    (17, 'month_cos',              216, 'Calendar (cyclical)',
     'Cosine encoding of month - paired with month_sin for cyclical encoding'),
    (18, 'quarter',                 11, 'Calendar',
     'Quarter of year - minimal marginal value once month is included'),
    (19, 'is_weekend',               0, 'Calendar',
     'Zero importance - confirmed no weekend effect in gaming hardware purchases'),
    (20, 'is_month_end',             0, 'Calendar',
     'Zero importance - no month-end purchasing pattern detected'),
    (21, 'is_q4',                    0, 'Calendar',
     'Zero importance - month feature already captures Q4 seasonality'),
    (22, 'regime',                   0, 'Change-point',
     'Zero importance - rolling features already encode which regime applies'),
]

category_colors = {
    'Rolling feature':    (LIGHT_BLUE,  MID_BLUE),
    'Product mix':        (LIGHT_GREEN, GREEN),
    'Calendar':           (LIGHT_AMBER, AMBER),
    'Calendar (cyclical)':(LIGHT_AMBER, AMBER),
    'Lag feature':        (LIGHT_GREY,  DARK_GREY),
    'Change-point':       (LIGHT_RED,   RED),
}

for i, (rank, feature, shap_val, category, interpretation) in enumerate(shap_data):
    row = 6 + i
    alt = i % 2 == 0
    write_data_row(ws4, row, [rank, feature, f'${shap_val:,}', category, interpretation], alt=alt)
    bg, fg = category_colors.get(category, (WHITE, DARK_GREY))
    c = ws4.cell(row=row, column=4)
    c.fill = fill(bg)
    c.font = Font(name=FONT, size=10, bold=True, color=fg)
    c.alignment = center()
    if shap_val == 0:
        for col in range(1, 6):
            ws4.cell(row=row, column=col).font = Font(name=FONT, size=10, color='AAAAAA')
    ws4.row_dimensions[row].height = 22

ws4.freeze_panes = 'A5'

# Sheet 5 - Recommendation
ws5 = wb.create_sheet('05 Recommendation')
ws5.sheet_view.showGridLines = False

title_block(ws5, 1, 'Model Recommendation and Next Steps', ncols=6)

section_label(ws5, 4, 'Model scorecard')
write_header_row(ws5, 5,
    ['Dimension', 'Prophet', 'LightGBM', 'Winner'],
    col_widths=[32, 28, 28, 14])

scorecard = [
    ('Training MAPE',              '25.1%',                  '0.83% (overfit)',         'Tie'),
    ('Honest CV MAPE',             'Not computed',           '11.6%',                   'LightGBM'),
    ('Shared test window MAPE',    f'{prophet_shared_mape:.0f}%', f'{lgbm_shared_mape:.0f}%', 'LightGBM'),
    ('Overfitting risk',           'Low',                    'High (small data)',        'Prophet'),
    ('Confidence intervals',       'Yes (built-in)',          'No',                      'Prophet'),
    ('Interpretability',           'Component decomposition','SHAP values',              'Tie'),
    ('Handles change-points',      'Yes (explicit)',          'Yes (via regime feature)', 'Tie'),
    ('Setup complexity',           'Low',                    'Medium',                  'Prophet'),
    ('Minimum data needed',        '2 years recommended',    '1 year minimum',          'LightGBM'),
    ('Adapts to sudden changes',   'No - projects trained trend', 'Yes - lag features update', 'LightGBM'),
]

winner_colors = {
    'Prophet':   (LIGHT_AMBER, AMBER),
    'LightGBM':  (LIGHT_GREEN, GREEN),
    'Tie':       (LIGHT_GREY,  DARK_GREY),
}

for i, (dim, p_val, l_val, winner) in enumerate(scorecard):
    row = 6 + i
    write_data_row(ws5, row, [dim, p_val, l_val, winner], alt=i % 2 == 0)
    bg, fg = winner_colors.get(winner, (WHITE, DARK_GREY))
    c = ws5.cell(row=row, column=4)
    c.fill = fill(bg)
    c.font = Font(name=FONT, size=10, bold=True, color=fg)
    c.alignment = center()
    ws5.row_dimensions[row].height = 20

offset = 6 + len(scorecard) + 2
section_label(ws5, offset, 'Recommendations')
write_header_row(ws5, offset+1,
    ['Scenario', 'Recommended Model', 'Rationale'],
    col_widths=[28, 20, 60])

recs = [
    ('This dataset (2 years, sparse)',
     'Prophet',
     'Lower overfitting risk, built-in confidence intervals, handles change-points explicitly'),
    ('Richer dataset (3+ years, stable)',
     'LightGBM',
     'Captures non-linear feature interactions; SHAP provides richer business insight'),
    ('Production deployment',
     'Ensemble (both)',
     'Weighted average of Prophet and LightGBM typically reduces individual model error'),
    ('Immediate next step',
     'Rebuild with richer data',
     'Find or generate a 3-5 year dataset - current data insufficient for reliable forecasting'),
]

rec_colors = {
    'Prophet':                (LIGHT_AMBER, AMBER),
    'LightGBM':               (LIGHT_GREEN, GREEN),
    'Ensemble (both)':        (LIGHT_BLUE,  MID_BLUE),
    'Rebuild with richer data':(LIGHT_RED,   RED),
}

for i, (scenario, model, rationale) in enumerate(recs):
    row = offset + 2 + i
    write_data_row(ws5, row, [scenario, model, rationale], alt=i % 2 == 0)
    bg, fg = rec_colors.get(model, (WHITE, DARK_GREY))
    c = ws5.cell(row=row, column=2)
    c.fill = fill(bg)
    c.font = Font(name=FONT, size=10, bold=True, color=fg)
    c.alignment = center()
    ws5.row_dimensions[row].height = 28

ws5.freeze_panes = 'A5'

wb.save(OUTPUT_PATH)
print(f'\nReport saved to {OUTPUT_PATH}')
print(f'Sheets: {wb.sheetnames}')
